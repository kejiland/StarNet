# -*- coding: utf-8 -*-
"""Excel 导出模块：生成带表头样式、自动换行、参考图（链接或嵌入图片）的产品信息表。

- 表头按“数据类别”切换：常规 / 京东 / 淘宝 / 拼多多；
- 京东表头：序号、商品名称、商品详情、商品参数、商品价格、参考图；
- 默认不下载图片，参考图直接写入图片 URL（最多 3 个、换行分隔）；
  若本地已下载图片（_image_path）则嵌入对应列。
"""

import math
import os
import time
from concurrent.futures import ThreadPoolExecutor
from io import BytesIO

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

try:
    from PIL import Image as PILImage
except Exception:  # 未安装 Pillow 时图片仅以路径文本展示
    PILImage = None

CATEGORIES = ["常规", "京东", "淘宝", "拼多多"]

# 每个数据类别对应的表头（第一列固定为“序号”）
CATEGORY_HEADERS = {
    "常规": ["序号", "产品名称", "用途", "设备介绍", "性能特点", "设备参数", "参考图片"],
    "京东": ["序号", "商品名称", "商品详情", "商品参数", "商品价格", "参考图"],
    "淘宝": ["序号", "产品名称", "用途", "设备介绍", "性能特点", "设备参数", "参考图片"],
    "拼多多": ["序号", "产品名称", "用途", "设备介绍", "性能特点", "设备参数", "参考图片"],
}

CATEGORY_WIDTHS = {
    "常规": [6, 24, 28, 40, 40, 40, 24],
    "京东": [6, 28, 42, 44, 14, 26],
    "淘宝": [6, 24, 28, 40, 40, 40, 24],
    "拼多多": [6, 24, 28, 40, 40, 40, 24],
}

# 跨类别取值兜底：语义等价字段
_FIELD_SYNONYMS = {
    "产品名称": ["产品名称", "商品名称"],
    "商品名称": ["商品名称", "产品名称"],
    "用途": ["用途"],
    "设备介绍": ["设备介绍", "商品详情", "商品介绍"],
    "商品详情": ["商品详情", "设备介绍", "商品介绍"],
    "性能特点": ["性能特点", "商品卖点", "商品特点"],
    "设备参数": ["设备参数", "商品参数", "规格参数"],
    "商品参数": ["商品参数", "设备参数", "规格参数"],
    "商品价格": ["商品价格", "价格"],
    "参考图片": ["参考图片", "参考图"],
    "参考图": ["参考图", "参考图片"],
}


def _norm_category(cat):
    """把类别写法归一化（兼容旧的“京东类/普通类”）。"""
    cat = str(cat or "").strip()
    if cat == "京东类":
        return "京东"
    if cat == "普通类" or cat == "":
        return "常规"
    return cat if cat in CATEGORIES else "常规"


def _detect_category(rows):
    """未指定类别时，从行数据中推断（全部行同类别则使用该类别，否则用常规）。"""
    cats = {_norm_category(r.get("_data_category") or r.get("数据类别") or "常规")
            for r in (rows or [])}
    if len(cats) == 1:
        return cats.pop()
    return "常规"


def _row_value(row, field):
    """按当前表头字段取值；字段不存在时用语义等价字段兜底。"""
    for k in _FIELD_SYNONYMS.get(field, [field]):
        v = row.get(k)
        if str(v or "").strip():
            return v
    return ""


def _build_image(img_path, image_w, image_h):
    """构造嵌入 Excel 的图片对象。

    - 小图（不超过展示框）直接引用原文件，不做任何解码缩放，速度最快；
    - 大图才解码缩放，避免对每张几 MB 的图片做无谓处理。
    """
    if PILImage is None:
        raise RuntimeError("缺少 Pillow")
    with PILImage.open(img_path) as im:
        w, h = im.size
    if w <= image_w and h <= image_h:
        xl = XLImage(img_path)
        return xl, w, h
    img = PILImage.open(img_path)
    img.thumbnail((image_w, image_h))
    buf = BytesIO()
    img.convert("RGB").save(buf, format="PNG")
    buf.seek(0)
    xl = XLImage(buf)
    return xl, img.width, img.height


def _prep_image(args):
    """多线程预处理单张图片，返回 (行号, (XLImage,宽,高) 或 None, 异常或 None)。"""
    idx, img_path, image_w, image_h = args
    try:
        return idx, _build_image(img_path, image_w, image_h), None
    except Exception as exc:
        return idx, None, exc


def _estimate_row_height(values, widths, has_img, image_h):
    max_lines = 1
    for v, w in zip(values, widths[1:]):
        text = str(v or "")
        if not text:
            continue
        per_line = max(4, int(w * 0.55))
        lines = 0
        for seg in text.split("\n"):
            lines += max(1, math.ceil(len(seg) / per_line))
        max_lines = max(max_lines, lines)
    height = 16 * max_lines + 10
    if has_img:
        height = max(height, image_h + 14)
    return min(height, 409)


def export_excel(rows, path, cfg, on_log=None, category=None):
    """把产品信息行导出为 Excel，返回输出路径。

    - category：数据类别（常规/京东/淘宝/拼多多），决定表头；为空时自动推断。
    """
    def log(msg):
        if on_log:
            on_log(msg)

    cat = _norm_category(category) if category else _detect_category(rows)
    headers = CATEGORY_HEADERS[cat]
    widths = CATEGORY_WIDTHS[cat]
    n_cols = len(headers)
    img_col = n_cols            # 参考图/参考图列 = 最后一列（Excel 列号从 1 开始）
    img_letter = get_column_letter(img_col)

    wb = Workbook()
    ws = wb.active
    ws.title = "产品信息"

    header_font = Font(name="宋体", size=14, bold=True)
    body_font = Font(name="宋体", size=11)

    for col, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = w

    for col in range(1, n_cols + 1):
        for rc in range(2, max(2, len(rows) + 2)):
            ws.cell(row=rc, column=col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    image_w = int(cfg["export"].get("image_width_px", 130))
    image_h = int(cfg["export"].get("image_height_px", 95))

    # 并行预处理图片（解码/缩放是导出中最耗时的部分）
    image_cache = {}
    image_errors = {}
    img_paths = {}
    for idx, row in enumerate(rows, start=2):
        p = row.get("_image_path") or ""
        if not p:
            p = _row_value(row, headers[-1])
        if p and os.path.isfile(p):
            img_paths[idx] = p
    if img_paths:
        workers = min(8, max(2, os.cpu_count() or 2))
        with ThreadPoolExecutor(max_workers=workers) as pool:
            for idx, result, err in pool.map(
                    _prep_image,
                    [(idx, p, image_w, image_h) for idx, p in img_paths.items()]):
                if err is None:
                    image_cache[idx] = result
                else:
                    image_errors[idx] = err

    total = len(rows)
    for idx, row in enumerate(rows, start=2):
        if total <= 20 or (idx - 1) % 10 == 0 or idx - 1 == total:
            log(f"正在导出第 {idx - 1}/{total} 行…")
        values = [idx - 1] + [_row_value(row, h) for h in headers[1:]]
        for col, v in enumerate(values, start=1):
            cell = ws.cell(row=idx, column=col, value=v)
            cell.font = body_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        has_img = False
        img_path = img_paths.get(idx, "")
        if idx in image_cache:
            xl_img, iw, ih = image_cache[idx]
            ws.add_image(xl_img, f"{img_letter}{idx}")
            has_img = True
        elif idx in image_errors:
            log(f"第 {idx - 1} 条图片嵌入失败（将显示路径）：{image_errors[idx]}")

        if not has_img:
            # 没有本地图片时，直接写入参考图片链接文本（可有多行）
            img_text = row.get("_image_path") or _row_value(row, headers[-1]) or ""
            cell = ws.cell(row=idx, column=img_col, value=img_text)
            cell.font = body_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ws.row_dimensions[idx].height = _estimate_row_height(values, widths, has_img, image_h)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{img_letter}{ws.max_row}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    os.makedirs(os.path.dirname(os.path.abspath(path)) or ".", exist_ok=True)
    t0 = time.time()
    wb.save(path)
    log(f"Excel 导出完成（数据类别：{cat}），共 {len(rows)} 条，用时 {time.time() - t0:.1f} 秒")
    return path
